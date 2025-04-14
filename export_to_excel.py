import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import os

def get_connection():
    conn = sqlite3.connect("scraper_data.db")
    return conn

def export_pdf_data(workbook):
    conn = get_connection()
    query = "SELECT id, region, pdf_title, pdf_link FROM pdf_data"
    df = pd.read_sql_query(query, conn)
    
    ws = workbook.create_sheet("PDF Data")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1EFFF", end_color="E1EFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 50
    
    conn.close()
    return workbook

def export_tariffs_data(workbook):
    conn = get_connection()
    query = "SELECT id, area, country, charge_type, port, currency FROM tariffs"
    df = pd.read_sql_query(query, conn)
    
    ws = workbook.create_sheet("Tariffs")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1EFFF", end_color="E1EFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    conn.close()
    return workbook

def export_container_types_data(workbook):
    conn = get_connection()
    query = """
    SELECT ct.id, t.area, t.country, t.port, 
           ct.type, ct.size, ct.free_time_days, ct.free_time_day_type,
           ct.detention_days, ct.detention_day_type, ct.detention_rate
    FROM container_types ct
    JOIN tariffs t ON ct.tariff_id = t.id
    """
    df = pd.read_sql_query(query, conn)
    
    ws = workbook.create_sheet("Container Types")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1EFFF", end_color="E1EFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    conn.close()
    return workbook

def export_rate_tiers_data(workbook):
    conn = get_connection()
    query = """
    SELECT rt.id, t.area, t.country, t.port, 
           ct.type, ct.size, 
           rt.tier_name, rt.start_day, rt.end_day, rt.rate, rt.rate_unit
    FROM rate_tiers rt
    JOIN container_types ct ON rt.container_type_id = ct.id
    JOIN tariffs t ON ct.tariff_id = t.id
    """
    df = pd.read_sql_query(query, conn)
    
    ws = workbook.create_sheet("Rate Tiers")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1EFFF", end_color="E1EFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    conn.close()
    return workbook

def create_comprehensive_view(workbook):
    conn = get_connection()
    query = """
    SELECT 
        t.area, t.country, t.charge_type, t.port, t.currency,
        ct.type as container_type, ct.size as container_size, 
        ct.free_time_days, ct.free_time_day_type,
        ct.detention_days, ct.detention_day_type, ct.detention_rate,
        rt.tier_name, rt.start_day, rt.end_day, rt.rate, rt.rate_unit
    FROM tariffs t
    JOIN container_types ct ON t.id = ct.tariff_id
    JOIN rate_tiers rt ON ct.id = rt.container_type_id
    ORDER BY t.area, t.country, t.port, ct.type, ct.size, rt.start_day
    """
    df = pd.read_sql_query(query, conn)
    
    ws = workbook.create_sheet("Comprehensive View")
    
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E1EFFF", end_color="E1EFFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    conn.close()
    return workbook

def export_db_to_excel():
    wb = Workbook()
    
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    wb = export_pdf_data(wb)
    wb = export_tariffs_data(wb)
    wb = export_container_types_data(wb)
    wb = export_rate_tiers_data(wb)
    wb = create_comprehensive_view(wb)
    
    filename = "scraper_data_export.xlsx"
    wb.save(filename)
    
    file_path = os.path.abspath(filename)
    print(f"Data exported successfully to: {file_path}")
    
    return file_path

if __name__ == "__main__":
    export_db_to_excel()